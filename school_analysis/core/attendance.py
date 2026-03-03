import streamlit as st
import pandas as pd
import datetime
import os
from openpyxl import load_workbook


def attendance_widget(df: pd.DataFrame):
    st.markdown("### 🚪 Отмечаем приход учеников")

    # --- Проверки данных ---
    required = {"class", "student"}
    missing = required - set(df.columns)
    if missing:
        st.error(f"В данных не хватает колонок: {missing}")
        return

    # --- Получаем классы и учеников ---
    classes = sorted(df["class"].dropna().unique())
    if not classes:
        st.info("Нет классов в данных.")
        return

    selected_class = st.selectbox("Выберите класс:", classes, key="attendance_class")
    students = sorted(df[df["class"] == selected_class]["student"].dropna().unique())

    if not students:
        st.info("В выбранном классе нет учеников.")
        return

    # --- Цветовые схемы ---
    color_cycle = {"gray": "green", "green": "yellow", "yellow": "red", "red": "gray"}
    bg_map = {
        "gray": "#dee2e6",     # серый
        "green": "#51cf66",    # зелёный
        "yellow": "#fcc419",   # жёлтый
        "red": "#ff6b6b",      # красный
    }
    text_map = {
        "gray": "#212529",
        "green": "#0b3d0b",
        "yellow": "#7f4f00",
        "red": "#7d0000",
    }
    emoji_map = {
        "gray": "⬜",
        "green": "🟩",
        "yellow": "🟨",
        "red": "🟥",
    }

    # ============================================================
    # ✅ Самовосстановление session_state под текущие данные
    # ============================================================
    if "attendance_status" not in st.session_state:
        st.session_state["attendance_status"] = {}

    if selected_class not in st.session_state["attendance_status"]:
        st.session_state["attendance_status"][selected_class] = {}

    # добавляем новых учеников
    for s in students:
        if s not in st.session_state["attendance_status"][selected_class]:
            st.session_state["attendance_status"][selected_class][s] = "gray"

    # (опционально) чистим тех, кого больше нет в текущем классе
    current_set = set(students)
    old_keys = list(st.session_state["attendance_status"][selected_class].keys())
    for s in old_keys:
        if s not in current_set:
            del st.session_state["attendance_status"][selected_class][s]

    # --- Плитки учеников ---
    st.markdown("#### 👩‍🏫 Отметьте статус учеников:")

    n_cols = 4 if len(students) > 12 else 3
    cols = st.columns(n_cols)

    for i, student in enumerate(students):
        col = cols[i % n_cols]

        # здесь больше не будет KeyError
        color = st.session_state["attendance_status"][selected_class].get(student, "gray")
        bg_color = bg_map[color]
        text_color = text_map[color]
        emoji = emoji_map[color]

        # CSS-трюк для цветной кнопки
        button_style = f"""
            <style>
            div[data-testid="stButton"][data-key="btn_{selected_class}_{i}"] button {{
                background-color: {bg_color};
                color: {text_color};
                font-weight: 600;
                border: 2px solid #adb5bd;
                border-radius: 10px;
                width: 100%;
                height: 60px;
                text-align: center;
                transition: transform 0.05s ease-in-out;
            }}
            div[data-testid="stButton"][data-key="btn_{selected_class}_{i}"] button:hover {{
                transform: scale(1.03);
            }}
            </style>
        """
        st.markdown(button_style, unsafe_allow_html=True)

        with col:
            if st.button(f"{student} {emoji}", key=f"btn_{selected_class}_{i}"):
                new_color = color_cycle[color]
                st.session_state["attendance_status"][selected_class][student] = new_color
                st.rerun()

    # --- Управление и журнал ---
    st.markdown("---")
    c1, c2, c3 = st.columns(3)

    with c1:
        if st.button("✅ Все пришли"):
            for s in students:
                st.session_state["attendance_status"][selected_class][s] = "green"
            st.rerun()

    with c2:
        if st.button("🔄 Сбросить"):
            for s in students:
                st.session_state["attendance_status"][selected_class][s] = "gray"
            st.rerun()

    # --- Данные для выгрузки/сохранения ---
    today = datetime.date.today().isoformat()
    data_out = pd.DataFrame(
        [
            {
                "date": today,
                "class": selected_class,
                "student": s,
                "status": st.session_state["attendance_status"][selected_class].get(s, "gray"),
                "timestamp": datetime.datetime.now().strftime("%H:%M:%S"),
            }
            for s in students
        ]
    )

    with c3:
        st.download_button(
            "⬇️ Скачать CSV",
            data=data_out.to_csv(index=False).encode("utf-8"),
            file_name=f"attendance_{selected_class}_{today}.csv",
            mime="text/csv",
        )

    # --- Сохранение в журнал ---
    log_path = "attendance_log.xlsx"
    if st.button("💾 Сохранить в журнал"):
        if not os.path.exists(log_path):
            data_out.to_excel(log_path, index=False)
        else:
            book = load_workbook(log_path)
            with pd.ExcelWriter(log_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                startrow = book.active.max_row
                data_out.to_excel(writer, index=False, header=False, startrow=startrow)
        st.success(f"Отметки сохранены в {log_path}")
